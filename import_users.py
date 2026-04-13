# -*- coding: utf-8 -*-
import openpyxl
import subprocess

MYSQL = r'C:\Program Files\MySQL\MySQL Server 8.4\bin\mysql.exe'

dept_name_to_id = {
    '富泉工贸': 100,
    '技术支持': 101,
    '公司领导': 102,
    '研发部门': 103,
    '行政部': 104,
    '测试部门': 105,
    '业务部': 106,
    '生产部': 107,
    '服装生产部': 108,
    '毛衣生产部': 109,
    '技术部': 110,
    '服装技术部': 111,
    '毛衣技术部': 112,
    '前道车间': 113,
    '后道车间': 114,
    '仓库': 115,
    '采购部': 116,
    '财务部': 117,
    '单证': 118,
}

sex_map = {'男': '0', '女': '1'}
status_map = {'正常': '0', '停用': '1'}

filepath = r'C:\Users\91306\Downloads\user_1775444860689.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb['用户数据']

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
col_map = {h: i for i, h in enumerate(headers)}

users = []
not_found_depts = set()
for row in range(2, ws.max_row + 1):
    def get_val(name):
        idx = col_map.get(name)
        if idx is None:
            return ''
        v = ws.cell(row=row, column=idx + 1).value
        return str(v).strip() if v else ''

    login_name = get_val('登录名称')
    nick_name = get_val('用户名称')
    dept_name = get_val('部门名称')
    email = get_val('用户邮箱')
    phone = get_val('手机号码')
    sex_raw = get_val('用户性别')
    status_raw = get_val('帐号状态')

    if not login_name:
        continue

    sex = sex_map.get(sex_raw, '2')
    status = status_map.get(status_raw, '0')
    dept_id = dept_name_to_id.get(dept_name)

    if dept_id is None:
        not_found_depts.add(dept_name)
        dept_id = 100

    users.append({
        'login_name': login_name,
        'nick_name': nick_name,
        'email': email,
        'phone': phone,
        'sex': sex,
        'status': status,
        'dept_id': dept_id,
        'dept_name': dept_name,
    })

print(f"共 {len(users)} 条用户")
if not_found_depts:
    print("✅ 所有部门都匹配成功!")
else:
    print(f"⚠️ 未匹配的部门: {not_found_depts}")

sql_parts = []
for u in users:
    login = u['login_name'].replace("'", "''")
    nick = u['nick_name'].replace("'", "''")
    email_val = (u['email'] or '').replace("'", "''")
    phone_val = (u['phone'] or '').replace("'", "''")

    sql_parts.append(
        f"INSERT INTO sys_user (user_name, nick_name, email, phonenumber, sex, avatar, password, dept_id, status, del_flag, create_by, create_time) "
        f"VALUES ('{login}', '{nick}', '{email_val}', '{phone_val}', '{u['sex']}', '', "
        f"'$2a$10$7JB720yubVSZvUI0rEqK/.VqGOZTH.ulu33dHOiBE8ByOhJIrdAu2', {u['dept_id']}, '{u['status']}', '0', 'admin', NOW()) "
        f"ON DUPLICATE KEY UPDATE nick_name='{nick}', email='{email_val}', phonenumber='{phone_val}', "
        f"sex='{u['sex']}', dept_id={u['dept_id']}, status='{u['status']}', update_by='admin', update_time=NOW();"
    )

full_sql = "\n".join(sql_parts)
proc = subprocess.Popen([MYSQL, '-u', 'root', '--default-character-set=utf8mb4', 'ry_vue'],
                       stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out, err = proc.communicate(input=full_sql.encode('utf-8'))
print("✅ SQL执行成功!" if proc.returncode == 0 else f"❌ 错误: {err.decode('utf-8')[:300]}")

verify = """
SELECT d.dept_name AS 部门, COUNT(u.user_id) AS 人数,
       GROUP_CONCAT(u.user_name ORDER BY u.user_id SEPARATOR ', ') AS 用户列表
FROM sys_dept d
LEFT JOIN sys_user u ON u.dept_id = d.dept_id AND u.del_flag = '0'
WHERE d.del_flag = '0' AND d.parent_id != 0
GROUP BY d.dept_id, d.dept_name
ORDER BY d.ancestors, d.order_num;
"""
proc2 = subprocess.Popen([MYSQL, '-u', 'root', '--default-character-set=utf8mb4', 'ry_vue'],
                        stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out2, _ = proc2.communicate(input=verify.encode('utf-8'))
print(f"\n=== 部门-用户分配 ===\n{out2.decode('utf-8').strip()}")
