# -*- coding: utf-8 -*-
import openpyxl
import subprocess

MYSQL = r'C:\Program Files\MySQL\MySQL Server 8.4\bin\mysql.exe'

dept_name_to_id = {
    '富泉工贸': 100, '技术支持': 101, '公司领导': 102, '研发部门': 103,
    '行政部': 104, '测试部门': 105, '业务部': 106, '生产部': 107,
    '服装生产部': 108, '毛衣生产部': 109, '技术部': 110, '服装技术部': 111,
    '毛衣技术部': 112, '前道车间': 113, '后道车间': 114, '仓库': 115,
    '采购部': 116, '财务部': 117, '单证': 118,
}
sex_map = {'男': '0', '女': '1'}
status_map = {'正常': '0', '停用': '1'}

filepath = r'C:\Users\91306\Downloads\user_1775444860689.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb['用户数据']
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
col_map = {h: i for i, h in enumerate(headers)}

def gv(row, name):
    idx = col_map.get(name)
    if idx is None: return ''
    v = ws.cell(row=row, column=idx+1).value
    return str(v).strip() if v else ''

updates = []
inserts = []
for row in range(2, ws.max_row + 1):
    login = gv(row, '登录名称')
    nick = gv(row, '用户名称')
    dept_name = gv(row, '部门名称')
    email = gv(row, '用户邮箱')
    phone = gv(row, '手机号码')
    sex_raw = gv(row, '用户性别')
    status_raw = gv(row, '帐号状态')
    if not login: continue

    sex = sex_map.get(sex_raw, '2')
    status = status_map.get(status_raw, '0')
    dept_id = dept_name_to_id.get(dept_name, 100)

    login_esc = login.replace("'", "''")
    nick_esc = nick.replace("'", "''")
    email_esc = (email or '').replace("'", "''")
    phone_esc = (phone or '').replace("'", "''")

    updates.append(
        f"UPDATE sys_user SET nick_name='{nick_esc}', email='{email_esc}', "
        f"phonenumber='{phone_esc}', sex='{sex}', dept_id={dept_id}, "
        f"status='{status}', update_by='admin', update_time=NOW() "
        f"WHERE user_name='{login_esc}';"
    )

full_sql = "\n".join(updates)
proc = subprocess.Popen([MYSQL, '-u', 'root', '--default-character-set=utf8mb4', 'ry_vue'],
                       stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out, err = proc.communicate(input=full_sql.encode('utf-8'))
print(f"✅ UPDATE 执行完成!" if proc.returncode == 0 else f"❌ {err.decode('utf-8')[:300]}")

verify = """
SELECT d.dept_name AS 部门, COUNT(u.user_id) AS 人数,
       GROUP_CONCAT(u.user_name ORDER BY u.user_id SEPARATOR ', ') AS 用户
FROM sys_dept d
LEFT JOIN sys_user u ON u.dept_id = d.dept_id AND u.del_flag='0'
WHERE d.del_flag='0' AND d.parent_id != 0
GROUP BY d.dept_id, d.dept_name
HAVING COUNT(u.user_id) > 0
ORDER BY d.ancestors, d.order_num;
"""
proc2 = subprocess.Popen([MYSQL, '-u', 'root', '--default-character-set=utf8mb4', 'ry_vue'],
                        stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out2, _ = proc2.communicate(input=verify.encode('utf-8'))
print(f"\n=== 部门-用户分配 ===\n{out2.decode('utf-8').strip()}")

total_sql = "SELECT COUNT(*) as total, SUM(CASE WHEN dept_id=100 THEN 1 ELSE 0 END) as root_dept FROM sys_user WHERE del_flag='0';"
proc3 = subprocess.Popen([MYSQL, '-u', 'root', '--default-character-set=utf8mb4', 'ry_vue'],
                        stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out3, _ = proc3.communicate(input=total_sql.encode('utf-8'))
print(f"\n{out3.decode('utf-8').strip()}")
