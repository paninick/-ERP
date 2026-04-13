import openpyxl
import os

filepath = r'C:\Users\91306\Downloads\user_1775444860689.xlsx'

if not os.path.exists(filepath):
    print(f"❌ 文件不存在: {filepath}")
    exit(1)

wb = openpyxl.load_workbook(filepath, data_only=True)
print(f"工作表: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"工作表: [{sheet_name}] - 共 {ws.max_row} 行 x {ws.max_column} 列")
    print(f"{'='*60}")
    
    print("\n--- 表头 ---")
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(val)
        print(f"  列{col}: {val}")
    
    print(f"\n--- 数据预览 (共 {ws.max_row-1} 条) ---")
    for row in range(2, min(ws.max_row + 1, 30)):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            row_data[headers[col-1]] = ws.cell(row=row, column=col).value
        print(f"  行{row}: {row_data}")
    
    if ws.max_row > 30:
        print(f"  ... 还有 {ws.max_row - 29} 条数据")
