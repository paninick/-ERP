import openpyxl

filepath = r'C:\Users\91306\Downloads\user_1775444860689.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb['用户数据']

dept_col_idx = None
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=c).value
    if val and '部门' in str(val):
        dept_col_idx = c
        print(f"部门列: 第{c}列, 表头='{val}'")
        break

print(f"\n前20行的部门列原始值 (col={dept_col_idx}):")
for row in range(2, min(22, ws.max_row + 1)):
    cell = ws.cell(row=row, column=dept_col_idx)
    val = cell.value
    val_type = type(val).__name__
    display = repr(val) if val else "None/空"
    login = ws.cell(row=row, column=2).value
    nick = ws.cell(row=row, column=3).value
    print(f"  行{row}: [{login}/{nick}] dept=({val_type}) {display}")

print(f"\n--- 检查是否有合并单元格 ---")
print(f"merged_cells: {ws.merged_cells.ranges}")
